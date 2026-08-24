#!/usr/bin/env python3
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl não instalado. Tentando instalar...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# Carregar planilha
wb = openpyxl.load_workbook("Apresent_1ºSemestre26.xlsx")

# Extrair dados de Liquidez_25 e Liquidez_26
data = {}
for sheet_name in ["Liquidez_25", "Liquidez_26"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
            if row[0]:  # Se primeira coluna não vazia
                sheet_data[str(row[0])] = row[1:]
        data[sheet_name] = sheet_data

print(json.dumps(data, indent=2, ensure_ascii=False))
