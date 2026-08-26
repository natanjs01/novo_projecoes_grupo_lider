from pathlib import Path
import json
import re

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "data.xlsx"
SLIDES = ROOT / "site" / "public" / "slides"
MONTHS = {"janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4, "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12}


def label(name):
	match = re.match(r"^(L\d+) LIDER (.+)$", name)
	return f"{match.group(1)} {match.group(2).title()}" if match else name


def is_store(name):
	return isinstance(name, str) and re.match(r"^L\d{2} LIDER ", name) and not name.startswith("L47 ")


def js(value):
	return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def main():
	ws = load_workbook(WORKBOOK, data_only=True)["Export"]
	year = month = store = None
	stores = {2025: {}, 2026: {}}
	segments = {2025: {}, 2026: {}}
	for row in ws.iter_rows(min_row=2, values_only=True):
		current_year, current_month, name, segment = row[:4]
		revenue, gross, net = row[8] or 0, row[10] or 0, row[12] or 0
		if current_year in (2025, 2026):
			year, month, store = current_year, None, None
		if current_month in MONTHS and name == "Total":
			month, store = MONTHS[current_month], None
		if year in (2025, 2026) and month in range(1, 7) and is_store(name) and segment == "Total":
			store = name
			stores[year].setdefault(store, [0, 0, 0])
			stores[year][store][0] += revenue
			stores[year][store][1] += gross
			stores[year][store][2] += net
		if year in (2025, 2026) and month in range(1, 7) and store and segment not in (None, "Total"):
			segments[year].setdefault(segment, {}).setdefault(store, [0, 0, 0])
			segments[year][segment][store][0] += revenue
			segments[year][segment][store][1] += gross
			segments[year][segment][store][2] += net

	ordered = sorted(stores[2026], key=lambda item: stores[2026][item][0], reverse=True)
	labels = [label(item) for item in ordered]
	revenues = {year: [round(stores[year][item][0]) for item in ordered] for year in (2025, 2026)}
	metrics = {year: {item: {
		"mb": stores[year][item][1] / stores[year][item][0] * 100 if stores[year][item][0] else 0,
		"ml": stores[year][item][2] / stores[year][item][0] * 100 if stores[year][item][0] else 0,
	} for item in ordered} for year in (2025, 2026)}
	margin_order = sorted(ordered, key=lambda item: metrics[2026][item]["mb"], reverse=True)
	segment_names = sorted(segments[2026], key=lambda item: sum(segments[2026][item][s][0] for s in segments[2026][item]), reverse=True)
	segment_totals = {year: {segment: [
		sum(segments[year].get(segment, {}).get(item, [0, 0, 0])[index] for item in ordered)
		for index in range(3)
	] for segment in segment_names} for year in (2025, 2026)}

	total_25, total_26 = sum(revenues[2025]), sum(revenues[2026])
	growth = (total_26 / total_25 - 1) * 100
	path = SLIDES / "14_faturamento_por_loja.html"
	content = path.read_text(encoding="utf-8")
	content = re.sub(r"var labels = \[.*?\];\n  var d25 = \[.*?\];\n  var d26 = \[.*?\];", f"var labels = {js(labels)};\n  var d25 = {js(revenues[2025])};\n  var d26 = {js(revenues[2026])};", content, count=1, flags=re.S)
	content = re.sub(r"(TOTAL 1S2025</p>\s*<p[^>]*>)R\$ [\d.]+ mi", rf"\g<1>R$ {total_25 / 1_000_000:.3f} mi", content, count=1)
	content = re.sub(r"(TOTAL 1S2026</p>\s*<p[^>]*>)R\$ [\d.]+ mi", rf"\g<1>R$ {total_26 / 1_000_000:.3f} mi", content, count=1)
	content = re.sub(r"[+-]\d+[,.]\d+%", f"{growth:+.1f}%".replace(".", ","), content, count=1)
	path.write_text(content, encoding="utf-8")

	path = SLIDES / "15_margens_e_rentabilidade.html"
	content = path.read_text(encoding="utf-8")
	raw = f"var labelsRaw = {js([label(item) for item in margin_order])};\n  var mb26Raw = {js([round(metrics[2026][item]['mb'], 2) for item in margin_order])};\n  var ml26Raw = {js([round(metrics[2026][item]['ml'], 2) for item in margin_order])};\n  var mb25Raw = {js([round(metrics[2025][item]['mb'], 2) for item in margin_order])};"
	content = re.sub(r"var labels = \[.*?\];\n  var mb26 = \[.*?\];\n  var ml26 = \[.*?\];\n  var mb25 = \[.*?\];", raw, content, count=1, flags=re.S)
	content = re.sub(r"var order = \[.*?\];", f"var order = {js(list(range(len(margin_order))))};", content, count=1)
	content = content.replace("Math.abs(mb25Raw[i])", "mb25Raw[i]")
	consolidated_mb = (sum(stores[2026][item][1] for item in ordered) / total_26) * 100
	consolidated_ml = (sum(stores[2026][item][2] for item in ordered) / total_26) * 100
	content = re.sub(r"Média MB [\d,]+% → [\d,]+% · Média ML [\d,]+% → [\d,]+%", f"Média MB {consolidated_mb:.1f}% → {consolidated_mb:.1f}% · Média ML {consolidated_ml:.1f}% → {consolidated_ml:.1f}%".replace(".", ","), content, count=1)
	path.write_text(content, encoding="utf-8")

	segment_values = {year: [round(segment_totals[year][segment][0]) for segment in segment_names] for year in (2025, 2026)}
	drilldown = {
		segment: [[label(item), round(values[0]), round(segments[2026].get(segment, {}).get(item, [0])[0])] for item, values in sorted(segments[2025].get(segment, {}).items(), key=lambda pair: pair[1][0], reverse=True) if item in ordered]
		for segment in segment_names
	}
	path = SLIDES / "16_faturamento_por_segmento.html"
	content = path.read_text(encoding="utf-8")
	content = re.sub(r"var order = \[.*?\];\n  var labelsRaw = \[.*?\];\n  var d25Raw = \[.*?\];\n  var d26Raw = \[.*?\];", f"var order = {js(list(range(len(segment_names))))};\n  var labelsRaw = {js(segment_names)};\n  var d25Raw = {js(segment_values[2025])};\n  var d26Raw = {js(segment_values[2026])};", content, count=1, flags=re.S)
	content = re.sub(r"var segmentDrilldown = \{.*?\n\n  var storeLabels", f"var segmentDrilldown = {js(drilldown)};\n\n  var storeLabels", content, count=1, flags=re.S)
	content = content.replace("8 segmentos", "9 segmentos")
	content = re.sub(r"(TOTAL 1S2025</p>\s*<p[^>]*>)R\$ [\d.]+ mi", rf"\g<1>R$ {total_25 / 1_000_000:.3f} mi", content, count=1)
	content = re.sub(r"(TOTAL 1S2026</p>\s*<p[^>]*>)R\$ [\d.]+ mi", rf"\g<1>R$ {total_26 / 1_000_000:.3f} mi", content, count=1)
	path.write_text(content, encoding="utf-8")

	margin_segments = sorted(segment_names, key=lambda segment: segment_totals[2026][segment][1] / segment_totals[2026][segment][0] if segment_totals[2026][segment][0] else 0, reverse=True)
	segment_mb = {year: [round(segment_totals[year][segment][1] / segment_totals[year][segment][0] * 100, 2) if segment_totals[year][segment][0] else 0 for segment in margin_segments] for year in (2025, 2026)}
	segment_ml = [round(segment_totals[2026][segment][2] / segment_totals[2026][segment][0] * 100, 2) if segment_totals[2026][segment][0] else 0 for segment in margin_segments]
	path = SLIDES / "17_margens_por_segmento.html"
	content = path.read_text(encoding="utf-8")
	content = re.sub(r"var order = \[.*?\];\n  var labelsRaw = \[.*?\];\n  var mb25Raw = \[.*?\];\n  var mb26Raw = \[.*?\];\n  var ml26Raw = \[.*?\];", f"var order = {js(list(range(len(margin_segments))))};\n  var labelsRaw = {js(margin_segments)};\n  var mb25Raw = {js(segment_mb[2025])};\n  var mb26Raw = {js(segment_mb[2026])};\n  var ml26Raw = {js(segment_ml)};", content, count=1, flags=re.S)
	content = content.replace("Math.abs(mb25Raw[i])", "mb25Raw[i]")
	content = content.replace("9 segmentos", "9 segmentos")
	path.write_text(content, encoding="utf-8")


if __name__ == "__main__":
	main()