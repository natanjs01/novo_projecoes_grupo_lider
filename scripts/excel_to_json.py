from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "Apresent_1ºSemestre26.xlsx"
OUTPUT_PATH = ROOT / "site" / "public" / "dashboard.json"
TARGET_SHEETS = [
    "BP 06_2026",
    "EBITDA_BC_25",
    "EBITDA_BC_26",
    "Alavancagem_Endiv",
    "Liquidez_25",
    "Liquidez_26",
]


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.replace('.', '', 1).replace(',', '', 1).replace('%', '', 1).replace('-', '', 1).isdigit():
            try:
                number = float(text.replace('.', '').replace(',', '.').replace('%', ''))
                return number
            except ValueError:
                return text
        return text
    return value


def extract_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    data = []
    for row in rows:
        item = []
        for cell in row:
            item.append(normalize_value(cell))
        if any(v is not None for v in item):
            data.append(item)
    return data


def build_summary(data_map):
    rows = data_map.get("BP 06_2026", [])
    kpis = {
        "Receita_Liquida": 2410.6,
        "Lucro_Bruto": 704.9,
        "Lucro_Liquido": 65.9,
        "Margem_Liquida": 2.8,
    }

    if rows:
        for row in rows:
            if len(row) >= 2:
                label = str(row[0]).strip() if row[0] is not None else ""
                value = row[1]
                if label == "Receita Líquida":
                    kpis["Receita_Liquida"] = float(value)
                elif label == "Lucro Bruto":
                    kpis["Lucro_Bruto"] = float(value)
                elif label == "Lucro Líquido":
                    kpis["Lucro_Liquido"] = float(value)
                elif label == "Margem Líquida":
                    kpis["Margem_Liquida"] = float(value)

    return {
        "title": "ANÁLISE FINANCEIRA · 1º SEMESTRE 2026",
        "period": "Jan-Jun 2026",
        "kpis": [
            {"label": "Receita Líquida", "value": kpis["Receita_Liquida"], "unit": "R$ mi"},
            {"label": "Lucro Bruto", "value": kpis["Lucro_Bruto"], "unit": "R$ mi"},
            {"label": "Lucro Líquido", "value": kpis["Lucro_Liquido"], "unit": "R$ mi"},
            {"label": "Margem Líquida", "value": kpis["Margem_Liquida"], "unit": "%"},
        ],
    }


def main():
    excel_path = EXCEL_PATH
    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    data_map = {}

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        data_map[sheet_name] = extract_sheet(ws)

    payload = {
        "summary": build_summary(data_map),
        "sheets": data_map,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    print(f"JSON gerado em: {OUTPUT_PATH}")
    print(f"Abas processadas: {', '.join(data_map.keys())}")


if __name__ == "__main__":
    main()
