import openpyxl
from pathlib import Path

xlsx_path = Path(r'\\10.15.4.252\Controladoria - Automação\Fábrica de sonhos\Natanael_BI_py\Apresentacao_grupo_lider_trimestral\nova_apresentacao\data.xlsx')
print('exists', xlsx_path.exists())
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print('sheets', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print('max_row', ws.max_row, 'max_col', ws.max_column)
for r in range(1, min(40, ws.max_row) + 1):
    vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column) + 1)]
    print(r, vals)
