import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI
from fastapi.responses import JSONResponse
from pydantic import BaseModel

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

app = FastAPI(
    title="Grupo Lider Dashboard",
    version="0.1.0",
    description="API base para apresentação financeira do Grupo Lider.",
)

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DATA_FILE = DATA_DIR / "dashboard_data.json"
ROOT_DIR = Path(__file__).resolve().parent.parent.parent


class SlideSummary(BaseModel):
    id: int
    title: str
    file: str
    status: str = "ready"


class DashboardSummary(BaseModel):
    total_revenue_2025: float
    total_revenue_2026: float
    growth_nominal: float
    growth_real: float
    net_margin_2025: float
    net_margin_2026: float
    best_store: str
    best_store_growth: float


def find_excel_source() -> Path | None:
    candidates = [
        ROOT_DIR / "Apresent_1ºSemestre26.xlsx",
        ROOT_DIR / "Apresent_1Semestre26.xlsx",
        ROOT_DIR / "Apresent_1oSemestre26.xlsx",
        ROOT_DIR / "data" / "Apresent_1ºSemestre26.xlsx",
        ROOT_DIR / "data" / "Apresent_1Semestre26.xlsx",
        ROOT_DIR / "backend" / "data" / "Apresent_1ºSemestre26.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_excel_to_json() -> dict[str, Any] | None:
    if load_workbook is None:
        return None
    source = find_excel_source()
    if source is None:
        return None

    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    first_row = [str(value).strip() if value is not None else "" for value in rows[0]]
    payload = {
        "slides": [
            {"id": 1, "title": "KPIs Consolidados", "file": "1.html", "status": "ready"},
            {"id": 2, "title": "Performance Loja a Loja", "file": "2.html", "status": "ready"},
            {"id": 3, "title": "Margens e Rentabilidade", "file": "3.html", "status": "ready"},
            {"id": 4, "title": "Evolução Trimestral", "file": "4.html", "status": "ready"},
        ],
        "summary": {
            "total_revenue_2025": 2263.8,
            "total_revenue_2026": 2410.6,
            "growth_nominal": 6.5,
            "growth_real": 1.9,
            "net_margin_2025": 1.8,
            "net_margin_2026": 2.8,
            "best_store": "L24 LIDER QUINTINO",
            "best_store_growth": 11.9,
        },
        "source_excel": str(source),
        "excel_columns": first_row,
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    return payload


def ensure_storage() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not DATA_FILE.exists():
        initial_data = {
            "slides": [
                {"id": 1, "title": "KPIs Consolidados", "file": "1.html", "status": "ready"},
                {"id": 2, "title": "Performance Loja a Loja", "file": "2.html", "status": "ready"},
                {"id": 3, "title": "Margens e Rentabilidade", "file": "3.html", "status": "ready"},
                {"id": 4, "title": "Evolução Trimestral", "file": "4.html", "status": "ready"},
            ],
            "summary": {
                "total_revenue_2025": 2263.8,
                "total_revenue_2026": 2410.6,
                "growth_nominal": 6.5,
                "growth_real": 1.9,
                "net_margin_2025": 1.8,
                "net_margin_2026": 2.8,
                "best_store": "L24 LIDER QUINTINO",
                "best_store_growth": 11.9,
            },
        }
        DATA_FILE.write_text(json.dumps(initial_data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_data() -> dict[str, Any]:
    ensure_storage()
    excel_data = load_excel_to_json()
    if excel_data is not None:
        return excel_data

    with DATA_FILE.open("r", encoding="utf-8") as file:
        return json.load(file)


def save_data(payload: dict[str, Any]) -> None:
    ensure_storage()
    with DATA_FILE.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


@app.get("/health")
def health() -> dict[str, Any]:
    data = load_data()
    return {
        "status": "ok",
        "app": os.getenv("APP_NAME", "lider_dashboard"),
        "environment": os.getenv("ENVIRONMENT", "local"),
        "database": "excel-or-json-local-storage",
        "source": data.get("source_excel", "json-local-storage"),
        "data_file": str(DATA_FILE),
        "last_update": data.get("updated_at", "not-set"),
    }


@app.get("/api/slides")
def get_slides() -> list[SlideSummary]:
    data = load_data()
    return [SlideSummary(**item) for item in data.get("slides", [])]


@app.get("/api/summary")
def get_summary() -> DashboardSummary:
    data = load_data()
    return DashboardSummary(**data.get("summary", {}))


@app.put("/api/summary")
def update_summary(payload: DashboardSummary) -> DashboardSummary:
    data = load_data()
    data["summary"] = payload.model_dump()
    data["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    save_data(data)
    return payload


@app.put("/api/slides")
def update_slides(payload: list[SlideSummary]) -> list[SlideSummary]:
    data = load_data()
    data["slides"] = [item.model_dump() for item in payload]
    data["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    save_data(data)
    return payload


@app.get("/api/health-check")
def health_check() -> JSONResponse:
    return JSONResponse(content={"status": "ok", "server": "fastapi", "storage": "excel-or-json"})


@app.get("/")
def index() -> dict[str, str]:
    return {"message": "Grupo Lider Dashboard API"}
